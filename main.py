import requests
import json
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timezone

# Remotive Jobs API endpoint
jobs_url = "https://remotive.com/api/remote-jobs"

# Filters
KEYWORDS = ["python", "ai", "data"]          # keyword filter (case-insensitive)
CATEGORY_FILTER = ""                         # Double check that this doesnt throw an error, example: "Software Development" (leave "" for no filter)
MAX_DAYS_OLD = 30                            # recent posting filter (keep jobs posted within last N days)

# Scoring Weights (should add up to 1.0)
WEIGHT_RECENCY = 0.50
WEIGHT_KEYWORDS = 0.40
WEIGHT_SALARY = 0.10

# Recency scoring rules:
# R = 1 when 0 days old
# R = 0.5 when 7 days old
# R = 0 when older than 7 days
RECENCY_CUTOFF_DAYS = 7
RECENCY_SLOPE_DENOM = 14  # makes R=0.5 at day 7 by dividing the 7/14 to get .5

OUTPUT_PATH = "./spreadsheets/remotive_jobs_scored.xlsx"
# ----------------------------
# HELPER FUNCTIONS
# ----------------------------
#note that using """ as a multiline string to make in code documentation

#Concerting all dates to timezone.UTC to be able to calculation recency 
def parse_iso_date(date_str):
    """ 
    Remotive publication_date is typically ISO 8601.
    Example: "2024-07-30T10:21:11+00:00"
    This returns a timezone-aware datetime, or None if it fails.
    """
    if date_str is None or date_str == "":
        return None

    try:
        dt = datetime.fromisoformat(date_str)
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt
    except ValueError:
        return None


def days_since(posted_dt):
    """
    Returns the number of whole days since posted_dt.
    If posted_dt is None, treat as old.
    """
    if posted_dt is None:
        return 9999 #treating no date as super old

    now = datetime.now(timezone.utc) # using the date now in same format as converted data from API
    delta = now - posted_dt.astimezone(timezone.utc)
    return delta.days #difference in days from posted to now


def keyword_match_count(text, keywords):
    """
    Counts how many keywords appear in the text (case-insensitive).
    Each keyword counts at most once.
    """
    if text is None:
        text = ""
#making case insensitive by converting keywords from API to lower case 
    text_lower = text.lower() #note on putting this outside the if loop so that it lowers all keyword and all job text to lower case once
    count = 0

    for kw in keywords:
        if kw.lower() in text_lower:
            count += 1

    return count

def recency_score(days_old):
    """
    Recency rules:
    - R = 1 when days_old = 0
    - R = 0.5 when days_old = 7
    - R = 0 when days_old > 7
    """
    if days_old > RECENCY_CUTOFF_DAYS:
        return 0

    r = 1 - (days_old / RECENCY_SLOPE_DENOM) 
    return max(0, r)


def keyword_score(match_count, total_keywords):
    """
    Normalizes keyword match count to 0-1.
    """
    if total_keywords == 0:
        return 0
    return match_count / total_keywords


def salary_score(salary_value):
    """
    Salary presence score: 1 if salary exists, else 0.
    """
    if salary_value is None:
        return 0
    if isinstance(salary_value, str) and salary_value.strip() == "":
        return 0
    return 1


def job_score(r, k, s):
    """
    Weighted score:
    JobScore = WEIGHT_RECENCY*R + WEIGHT_KEYWORDS*K + WEIGHT_SALARY*S
    """
    return (WEIGHT_RECENCY * r) + (WEIGHT_KEYWORDS * k) + (WEIGHT_SALARY * s)


def passes_filters(job, keywords, category_filter, max_days_old):
    """
    Returns True if job passes keyword, category, and recency filters.
    """
    title = job.get("title", "")
    description = job.get("description", "")
    category = job.get("category", "")
    pub_date_str = job.get("publication_date", "")

    text = f"{title} {description}"

    # Keyword filter (if keywords provided)
    if len(keywords) > 0:
        matches = keyword_match_count(text, keywords)
        if matches == 0:
            return False

    # Category filter (if set)
    if category_filter != "":
        if category != category_filter:
            return False

    # Recency filter (keep jobs within last max_days_old days)
    posted_dt = parse_iso_date(pub_date_str)
    d_old = days_since(posted_dt)
    if d_old > max_days_old:
        return False

    return True


# ----------------------------
# API CALL + JSON PARSING
# ----------------------------

response = requests.get(jobs_url)
print("STATUS CODE:", response.status_code)

data = json.loads(response.text)

# Remotive API typically returns jobs under the "jobs" key
jobs_list = data.get("jobs", [])
print("Total jobs pulled from API:", len(jobs_list))


# ----------------------------
# FILTER + SCORE
# ----------------------------

filtered_jobs = []

for job in jobs_list:
    if passes_filters(job, KEYWORDS, CATEGORY_FILTER, MAX_DAYS_OLD):
        title = job.get("title", "")
        company = job.get("company_name", "")
        category = job.get("category", "")
        pub_date_str = job.get("publication_date", "")
        url = job.get("url", "")
        salary = job.get("salary", "")
        description = job.get("description", "")

        posted_dt = parse_iso_date(pub_date_str)
        d_old = days_since(posted_dt)

        text = f"{title} {description}"
        match_count = keyword_match_count(text, KEYWORDS)

        r = recency_score(d_old)
        k = keyword_score(match_count, len(KEYWORDS))
        s = salary_score(salary)

        score = job_score(r, k, s)

        filtered_jobs.append({
            "title": title,
            "company": company,
            "category": category,
            "publication_date": pub_date_str,
            "days_old": d_old,
            "salary": salary,
            "keyword_match_count": match_count,
            "recency_score": r,
            "keyword_score": k,
            "salary_score": s,
            "job_score": score,
            "url": url
        })

print("Total jobs after filters:", len(filtered_jobs))


# Sort by job_score descending
filtered_jobs.sort(key=lambda x: x["job_score"], reverse=True)
# for each item x in the list, use its job score value as its sorting value

# ----------------------------
# EXCEL OUTPUT (OpenPyXL)
# ----------------------------

wb = Workbook()

# Jobs sheet
ws_jobs = wb.active
ws_jobs.title = "Jobs"

headers = [
    "Title",
    "Company",
    "Category",
    "Publication Date",
    "Days Since Posted",
    "Salary",
    "Keyword Match Count",
    "Recency Score (R)",
    "Keyword Score (K)",
    "Salary Score (S)",
    "Job Score",
    "Link"
]

for col, header in enumerate(headers, 1):
    ws_jobs.cell(row=1, column=col, value=header)

# Write job rows
for row, job in enumerate(filtered_jobs, 2):
    ws_jobs.cell(row=row, column=1, value=str(job["title"]))
    ws_jobs.cell(row=row, column=2, value=str(job["company"]))
    ws_jobs.cell(row=row, column=3, value=str(job["category"]))
    ws_jobs.cell(row=row, column=4, value=str(job["publication_date"]))
    ws_jobs.cell(row=row, column=5, value=job["days_old"])
    ws_jobs.cell(row=row, column=6, value=str(job["salary"]))
    ws_jobs.cell(row=row, column=7, value=job["keyword_match_count"])
    ws_jobs.cell(row=row, column=8, value=job["recency_score"])
    ws_jobs.cell(row=row, column=9, value=job["keyword_score"])
    ws_jobs.cell(row=row, column=10, value=job["salary_score"])
    ws_jobs.cell(row=row, column=11, value=job["job_score"])

    # Hyperlink
    link_cell = ws_jobs.cell(row=row, column=12, value="View Job")
    if job["url"] != "":
        link_cell.hyperlink = job["url"]
        link_cell.style = "Hyperlink"

# Column widths (simple)
for col in range(1, len(headers) + 1):
    col_letter = get_column_letter(col)
    ws_jobs.column_dimensions[col_letter].width = 20

ws_jobs.column_dimensions["A"].width = 45  # Title wider
ws_jobs.column_dimensions["L"].width = 15  # Link


# Summary sheet
ws_summary = wb.create_sheet("Summary")

ws_summary["A1"] = "Remote Job Finder Summary"

ws_summary["A3"] = "API Endpoint"
ws_summary["B3"] = jobs_url

ws_summary["A5"] = "Keywords"
ws_summary["B5"] = ", ".join(KEYWORDS)

ws_summary["A6"] = "Category Filter"
ws_summary["B6"] = CATEGORY_FILTER if CATEGORY_FILTER != "" else "(none)"

ws_summary["A7"] = "Max Days Old (Filter)"
ws_summary["B7"] = MAX_DAYS_OLD

ws_summary["A9"] = "Total Jobs Pulled"
ws_summary["B9"] = len(jobs_list)

ws_summary["A10"] = "Total Jobs After Filters"
ws_summary["B10"] = len(filtered_jobs)

ws_summary["A12"] = "Top Jobs"
ws_summary["A13"] = "Rank"
ws_summary["B13"] = "Title"
ws_summary["C13"] = "Company"
ws_summary["D13"] = "Job Score"
ws_summary["E13"] = "Link"

top_n = min(10, len(filtered_jobs))
for i in range(top_n):
    job = filtered_jobs[i]
    r = 14 + i  # start row for top jobs list

    ws_summary.cell(row=r, column=1, value=i + 1)
    ws_summary.cell(row=r, column=2, value=str(job["title"]))
    ws_summary.cell(row=r, column=3, value=str(job["company"]))
    ws_summary.cell(row=r, column=4, value=job["job_score"])

    link_cell = ws_summary.cell(row=r, column=5, value="View Job")
    if job["url"] != "":
        link_cell.hyperlink = job["url"]
        link_cell.style = "Hyperlink"

ws_summary.column_dimensions["A"].width = 10
ws_summary.column_dimensions["B"].width = 55
ws_summary.column_dimensions["C"].width = 25
ws_summary.column_dimensions["D"].width = 15
ws_summary.column_dimensions["E"].width = 15


# Save workbook
wb.save(OUTPUT_PATH)
print("Saved:", OUTPUT_PATH)